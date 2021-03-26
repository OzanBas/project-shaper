from PyQt5.QtWidgets import QFileDialog, QMessageBox
import openpyxl
import openpyxl.styles
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import numbers, Font, PatternFill, colors






def func_Single():
    
    #### FILE LOAD ####

    pathlist = QFileDialog.getOpenFileName()
    cancel_Fix = str(pathlist)
    if cancel_Fix.endswith("('', '')"):
        return
    
    path = ''.join(pathlist[0])
    print(path)
    
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    
    nameplate = path.split('/')[-1]
    
    
    #### PATH & NAME FORMAT ####
    XL = '.XLSX'
    xl = '.xlsx'
    if nameplate.endswith(XL):
        name = nameplate.replace(XL, '')
        path = path.replace(XL, '.XLSX')
    elif nameplate.endswith(xl):
        name = nameplate.replace(xl, '')
        path = path.replace(xl, '.xlsx')

    
    #### RESHAPE EXCEL ####
    ws.delete_cols(5, 9)
    ws.delete_cols(1)

    for row in ws['A1:A2']:
        for cell in row:
            cell.value = None

    for row in range(1, 2000):
        ws["C{}".format(row)].number_format = '#,###0.000'

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 12

        fontStyle = Font(size = "14", bold = True)
        ws.cell(row = 1, column = 1, value = name).font = fontStyle


    #### OUTPUT & SAVE ####
    wb.save(path)
    
    msg = QMessageBox()
    msg.setIcon(QMessageBox.Information)
    msg.setText('"' + name + '" excel dosyanız yeniden düzenlendi.')
    msg.exec_()
    
###########################################################################
def func_Multiple():

    from datetime import datetime
    import os
    import pathlib

    #### FILE LOAD & START LOOP FILES ####
    path = QFileDialog.getOpenFileNames()

    output = 0

       
    
    
    for i in path[0]:

        path = ''.join(i)
        wb = openpyxl.load_workbook(i)
        ws = wb.active


        #### PATH & NAME FORMAT ####
        nameplate = i.split('/')[-1]
        
        
        XL = '.XLSX'
        xl = '.xlsx'
        
        if nameplate.endswith(XL):
            name = nameplate.replace(XL, '')
            #i = i.replace(XL, 'r.XLSX')
        elif nameplate.endswith(xl):
            name = nameplate.replace(xl, '')
            #i = i.replace(xl, 'r.xlsx')
        
        
        #### EXCEL RESHAPE ####
        ws.delete_cols(5, 9)
        ws.delete_cols(1)

        for row in ws['A1:A2']:
            for cell in row:
                cell.value = None

        for row in range(1, 2000):
            ws["C{}".format(row)].number_format = '#,###0.000'

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 12

        fontStyle = Font(size = "14", bold = True)
        ws.cell(row = 1, column = 1, value = name).font = fontStyle


        #### OUTPUT & SAVE ####
        save_Name = datetime.now().strftime('Reshaped %d-%m-%Y')
        save_Dir = os.path.join(pathlib.Path.home()/'Desktop', save_Name)
        os.makedirs(save_Dir, exist_ok=True)


        save_Path = os.path.join(save_Dir, nameplate)
        wb.save(save_Path)
        
        output += 1
    if output >= 1:

        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText(str(output) + ' adet excel dosyası oluşturuldu')
        msg.exec_()
    
###########################################################################
def func_Clean():
    
    
    #### FILE LOAD & START LOOP FILES ####    
    path = QFileDialog.getOpenFileNames()
    output = 0
    
    for i in path[0]:
        path = ''.join(i)
        wb = openpyxl.load_workbook(i)
        ws = wb.active


        #### PATH & NAME FORMAT ####
        nameplate = i.split('/')[-1]
        XL = '.XLSX'
        xl = '.xlsx'
        
        if nameplate.endswith(XL):
            name = nameplate.replace(XL, '')
            i = i.replace(XL, 'Mutfak.XLSX')
        elif nameplate.endswith(xl):
            name = nameplate.replace(xl, '')
            i = i.replace(xl, 'Mutfak.xlsx')
        
        
        #### EXCEL RESHAPE ####
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))



        #ADDING BORDERS
        for s in range(1,ws.max_row):

            if not ws.cell(s,5).value == None:
                ws.cell(s,2).border = thin_border
                ws.cell(s,3).border = thin_border
                ws.cell(s,4).border = thin_border


        #DELETE UNNECESSARY COLUMNS
        ws.delete_cols(5, 9)
        ws.delete_cols(1)


        #DELETE FIRST 2 LINES
        for row in ws['A1:A2']:
            for cell in row:
                cell.value = None

        
        #CHANGE NUMBER FORMAT
        for row in range(ws.min_row, ws.max_row):
            ws["C{}".format(row)].number_format = '#,###0.000'


        # SETTING DIMENSIONS
        ws.column_dimensions['A'].width = 42
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 14

        fontStyle = Font(size = "18", bold = True)
        ws.cell(row = 1, column = 1, value = name).font = fontStyle





        #### OUTPUT & SAVE ####
        wb.save(i)
        output += 1
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText(str(output) + ' adet Mutfak Reçetesi oluşturuldu.')
        msg.exec_()
        
###########################################################################